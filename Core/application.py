import string
import random
from datetime import timedelta, datetime
import calendar
from django.utils.timezone import make_aware #type: ignore
import openpyxl, io #type: ignore
from django.apps import apps #type: ignore
from django.http import HttpResponse #type: ignore
from django.contrib.admin.utils import label_for_field #type: ignore

from .models import *

def generate_password(length=8):
    chars = string.ascii_letters + string.digits + "!@#$%^&*"
    return ''.join(random.choices(chars, k=length))

def generate_username(firstname, middlename, surname):
    # Step 1: Take initials
    initials = (firstname[:random.randint(1, 2)] + (middlename[:random.randint(1, 2)] if middlename else "") + surname[:random.randint(1, 3)]).lower()

    # Step 2: Ensure it's 3–4 letters before number
    initials = initials[:5]

    # Step 3: Append number until unique
    while True:
        num = ''.join(random.choices(string.digits, k=2))  # 2-digit number
        username = f"{initials}{num}"
        if not User.objects.filter(username=username).exists():
            return username

def get_or_create_module_by_name(module_input_name: str):
    # Normalize the input
    name_cleaned = module_input_name.strip().lower()

    # Map possible user inputs to canonical module name and abbreviation
    module_mapping = {
        '1': ('Module One', 'M1'),
        'one': ('Module One', 'M1'),
        'module one': ('Module One', 'M1'),
        'module 1': ('Module One', 'M1'),

        '2': ('Module Two', 'M2'),
        'two': ('Module Two', 'M2'),
        'module two': ('Module Two', 'M2'),
        'module 2': ('Module Two', 'M2'),

        '3': ('Module Three', 'M3'),
        'three': ('Module Three', 'M3'),
        'module three': ('Module Three', 'M3'),
        'module 3': ('Module Three', 'M3'),

        '4': ('Module Four', 'M4'),
        'four': ('Module Four', 'M4'),
        'module four': ('Module Four', 'M4'),
        'module 4': ('Module Four', 'M4'),

        '5': ('Module Five', 'M5'),
        'five': ('Module Five', 'M5'),
        'module five': ('Module Five', 'M5'),
        'module 5': ('Module Five', 'M5'),
    }

    # Resolve the intended canonical name/abbr
    module_data = module_mapping.get(name_cleaned)

    if not module_data:
        raise ValueError(f"Unrecognized module name: '{module_input_name}'")

    canonical_name, abbr = module_data

    # Try to find existing module
    module = Module.objects.filter(name__iexact=canonical_name).first()
    if not module:
        module = Module.objects.create(name=canonical_name, abbr=abbr)

    return module  # or return module.id if only ID is needed

def get_opening_date(year: int, month_name: str):
    month = list(calendar.month_name).index(month_name)
    d = datetime(year, month, 1)
    # Move to the second Monday
    mondays = [d + timedelta(days=i) for i in range(31) if (d + timedelta(days=i)).month == month and (d + timedelta(days=i)).weekday() == 0]
    return make_aware(mondays[1]) if len(mondays) > 1 else make_aware(mondays[0])

def get_closing_date(year: int, month_name: str):
    month = list(calendar.month_name).index(month_name)
    last_day = calendar.monthrange(year, month)[1]
    d = datetime(year, month, last_day)
    # Move backward to the last Friday
    while d.weekday() != 4:  # 4 = Friday
        d -= timedelta(days=1)
    return make_aware(d)

def download_school_template(request):
    # Create a new Excel workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove the default sheet

    # Loop through all models in all installed apps
    for model in apps.get_models():
        sheet_name = model.__name__[:31]  # Excel sheet name max 31 chars
        ws = wb.create_sheet(title=sheet_name)

        # Get field names
        field_names = [field.name for field in model._meta.get_fields() if not field.is_relation or field.many_to_one]

        # Write header row
        for col_num, field_name in enumerate(field_names, start=1):
            ws.cell(row=1, column=col_num, value=label_for_field(field_name, model))

    # Save workbook to memory
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # Create response for download
    response = HttpResponse(
        excel_file,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="school_template.xls"'
    return response

def is_student_user(user):
    """
    Returns True if the user belongs to a group named like 'student' 
    (case-insensitive, singular/plural).
    """
    student_groups = ["student", "students"]  # accepted variants
    user_groups = user.groups.values_list("name", flat=True)

    return any(g.lower() in student_groups for g in user_groups)

def is_rep_user(user):
    """
    Returns True if the user belongs to a group named like 'class rep' 
    (case-insensitive, singular/plural).
    """
    student_groups = ["class rep", "rep", 
                      "class representative", 
                      "representative", "reps", 
                      "class reps", "representatives", 
                      "class representatives",
                      "prefect", "class prefect", "prefects",
                      "class prefects"
    ]  # accepted variants
    user_groups = user.groups.values_list("name", flat=True)

    return any(g.lower() in student_groups for g in user_groups)

def is_staff_user(user):
    """
    Returns True if the user belongs to a group named like 'staff' 
    (case-insensitive, singular/plural).
    """
    staff_groups = ["staff", "staffs", "lecturer", "lecturers", "teacher", "teachers"]  # accepted variants
    user_groups = user.groups.values_list("name", flat=True)

    return any(g.lower() in staff_groups for g in user_groups)

def is_tutor_user(user):
    """
    Returns True if the user belongs to a group named like 'class tutor' 
    (case-insensitive, singular/plural).
    """
    staff_groups = ["class tutor", "tutor", 
                      "class tutors", 
                      "tutors", "class teacher", 
                      "class teachers"
    ]  # accepted variants
    user_groups = user.groups.values_list("name", flat=True)

    return any(g.lower() in staff_groups for g in user_groups)

def is_admin_user(user):
    """
    Returns True if the user belongs to a group named like 'admin' 
    (case-insensitive, singular/plural).
    """
    admin_groups = ["admin", "administrators", "user"]  # accepted variants
    user_groups = user.groups.values_list("name", flat=True)

    return any(g.lower() in admin_groups for g in user_groups)
